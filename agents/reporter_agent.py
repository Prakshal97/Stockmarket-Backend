"""
Reporter Agent — Generates styled Excel reports with STRICT SEGREGATION.

Sheet 1: Authorized Capital (15 columns — client spec)
Sheet 2: Other Announcements (7 columns)
"""
import io
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Market Data Helpers ───────────────────────────────────────────────────

def _prefetch_market_data(announcements: List[Dict]) -> dict:
    """Pre-fetch live CMP & Market Cap for all unique tickers."""
    try:
        from agents.market_data import batch_get_market_data, _clean_ticker
        tickers = []
        for ann in announcements:
            ai_data = ann.get("ai_data", {}) or {}
            t = ai_data.get("ticker") or ann.get("ticker") or ann.get("symbol", "")
            if t:
                tickers.append(t)
        if not tickers:
            return {}
        return batch_get_market_data(tickers)
    except Exception as e:
        print(f"WARNING: batch market data prefetch failed: {e}")
        return {}


def _resolve_cmp(ai_data: dict, mkt_cache: dict) -> str:
    """Return CMP as formatted string."""
    val = ai_data.get("cmp")
    if val is not None:
        try:
            return f"₹{float(val):,.2f}"
        except:
            return str(val)
    ticker = ai_data.get("ticker", "")
    if ticker:
        try:
            from agents.market_data import _clean_ticker
            clean = _clean_ticker(ticker)
            mkt = mkt_cache.get(clean, {})
            live_cmp = mkt.get("cmp")
            if live_cmp is not None:
                return f"₹{float(live_cmp):,.2f}"
        except:
            pass
    return "Unavailable"


def _resolve_market_cap(ai_data: dict, mkt_cache: dict) -> str:
    """Return Market Cap as formatted string."""
    val = ai_data.get("market_cap_cr")
    if val is not None:
        try:
            return f"₹{float(val):,.2f} Cr"
        except:
            return str(val)
    ticker = ai_data.get("ticker", "")
    if ticker:
        try:
            from agents.market_data import _clean_ticker
            clean = _clean_ticker(ticker)
            mkt = mkt_cache.get(clean, {})
            live_mcap = mkt.get("market_cap_cr")
            if live_mcap is not None:
                return f"₹{float(live_mcap):,.2f} Cr"
        except:
            pass
    return "Unavailable"


# ── Color Palette ─────────────────────────────────────────────────────────

COLORS = {
    "header_bg": "1E3A5F",
    "header_font": "FFFFFF",
    "title_bg": "0D1B2A",
    "positive_bg": "D5F5E3",
    "negative_bg": "FADBD8",
    "neutral_bg": "FEF9E7",
    "alt_row": "EBF5FB",
    "border": "BDC3C7",
    "auth_title_bg": "1B4332",
    "general_title_bg": "1E3A5F",
}

THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS["border"]),
    right=Side(style='thin', color=COLORS["border"]),
    top=Side(style='thin', color=COLORS["border"]),
    bottom=Side(style='thin', color=COLORS["border"])
)


def _fmt_currency(val) -> str:
    """Format INR value."""
    if val is None:
        return "Unavailable"
    try:
        val = float(val)
        crores = val / 1_00_00_000
        if crores >= 1:
            return f"₹{crores:,.2f} Cr"
        lakhs = val / 1_00_000
        if lakhs >= 1:
            return f"₹{lakhs:,.2f} L"
        return f"₹{val:,.0f}"
    except:
        return str(val) if val else "Unavailable"


# ══════════════════════════════════════════════════════════════════════════
# MAIN EXPORT: Segregated Report (Sheet 1: Auth Cap, Sheet 2: Others)
# ══════════════════════════════════════════════════════════════════════════

def generate_segregated_excel(auth_announcements: List[Dict], general_announcements: List[Dict]) -> bytes:
    """
    Generate a dual-section Excel report.
    Sheet 1: Authorized Capital (15 columns, client spec)
    Sheet 2: Other Announcements (7 columns)
    """
    all_anns = auth_announcements + general_announcements
    mkt_cache = _prefetch_market_data(all_anns)

    wb = Workbook()

    # ── Sheet 1: Authorized Capital ───────────────────────────────
    ws1 = wb.active
    ws1.title = "Authorized Capital"
    _write_auth_capital_sheet(ws1, auth_announcements, mkt_cache)

    # ── Sheet 2: Other Announcements ──────────────────────────────
    ws2 = wb.create_sheet("Other Announcements")
    _write_general_sheet(ws2, general_announcements)

    return _save_workbook_bytes(wb)


def generate_authorized_capital_excel(announcements: List[Dict]) -> bytes:
    """Generate Excel with ONLY authorized capital data (backward compat)."""
    mkt_cache = _prefetch_market_data(announcements)
    wb = Workbook()
    ws = wb.active
    ws.title = "Authorized Capital"
    _write_auth_capital_sheet(ws, announcements, mkt_cache)
    return _save_workbook_bytes(wb)


def generate_full_report_excel(announcements: List[Dict]) -> bytes:
    """Legacy: Generate report from mixed announcements (backward compat)."""
    mkt_cache = _prefetch_market_data(announcements)
    wb = Workbook()
    ws = wb.active
    ws.title = "Full AI Report"
    _write_general_sheet(ws, announcements)
    return _save_workbook_bytes(wb)


# ══════════════════════════════════════════════════════════════════════════
# Sheet Writers
# ══════════════════════════════════════════════════════════════════════════

def _write_auth_capital_sheet(ws, announcements: List[Dict], mkt_cache: dict):
    """
    Write Authorized Capital sheet with the exact 10 columns required by the client.
    """
    # ── Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"SECTION 1: AUTHORIZED CAPITAL — {datetime.now().strftime('%d %B %Y')}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLORS["header_font"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["auth_title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Headers (15 columns) ──────────────────────────────────────
    headers = [
        "Company", "Symbol", "Exchange", "Date", "Old Capital",
        "New Capital", "Increase Amount", "% Increase", "Source URL", "PDF URL"
    ]
    col_widths = [30, 14, 12, 14, 18, 18, 18, 12, 28, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 32

    # ── Data Rows ─────────────────────────────────────────────────
    if not announcements:
        ws.merge_cells("A3:J3")
        cell = ws["A3"]
        cell.value = "No Authorized Capital announcements found in the last 48 hours."
        cell.font = Font(name="Calibri", size=11, italic=True, color="888888")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A3"
        return

    for sr_no, ann in enumerate(announcements, 1):
        ai_data = ann.get("ai_data", {}) or {}
        auth_cap = ai_data.get("authorized_capital", {}) or {}
        auth_data_top = ann.get("auth_data", {}) or {}
        merged_auth = {**auth_cap}
        for key, val in auth_data_top.items():
            if val is not None and val != "Not Available":
                merged_auth[key] = val

        try:
            ann_date = datetime.fromisoformat(ann.get("announcement_date", "")).strftime("%d.%m.%Y")
        except:
            ann_date = ""

        old_cap = ann.get("old_capital_inr") or merged_auth.get("existing_auth_eq_cap_inr")
        new_cap = ann.get("new_capital_inr") or merged_auth.get("new_auth_eq_cap_inr")
        increase_amt = None
        pct_increase = ann.get("percentage_increase")
        if isinstance(old_cap, (int, float)) and old_cap < 1_00_000:
            old_cap = None
        if isinstance(new_cap, (int, float)) and new_cap < 1_00_000:
            new_cap = None
        if old_cap and new_cap and new_cap > old_cap:
            increase_amt = round(float(new_cap) - float(old_cap), 2)
        else:
            increase_amt = ann.get("increase_amount_inr") or merged_auth.get("proposed_increase_inr")
            if isinstance(increase_amt, (int, float)) and increase_amt < 1_00_000:
                increase_amt = None
        if old_cap and new_cap and new_cap <= old_cap:
            increase_amt = None
            pct_increase = None
        if pct_increase is None and old_cap and new_cap:
            try:
                pct_increase = round(((float(new_cap) - float(old_cap)) / float(old_cap)) * 100, 2)
            except Exception:
                pct_increase = None

        row_data = [
            ai_data.get("company_name") or ann.get("company_name", ""),
            ai_data.get("ticker") or ann.get("symbol") or ann.get("ticker", ""),
            ann.get("exchange", "NSE"),
            ann_date,
            _fmt_currency(old_cap),
            _fmt_currency(new_cap),
            _fmt_currency(increase_amt),
            f"{pct_increase:.2f}%" if pct_increase is not None else "Unavailable",
            ann.get("source_url", ""),
            ann.get("pdf_url", ""),
        ]

        row = sr_no + 2
        row_fill = PatternFill("solid", fgColor=COLORS["alt_row"]) if sr_no % 2 == 0 else None

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=9)
            cell.border = THIN_BORDER
            if col_idx in [1, 9, 10]:
                align = "left"
            else:
                align = "center"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row].height = 26

    ws.freeze_panes = "A3"


def _write_general_sheet(ws, announcements: List[Dict]):
    """
    Write Other Announcements sheet with 7 columns.
    """
    # ── Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"SECTION 2: OTHER ANNOUNCEMENTS — {datetime.now().strftime('%d %B %Y')}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLORS["header_font"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["general_title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Headers (7 columns) ───────────────────────────────────────
    headers = [
        "Company", "Date", "Category", "Sentiment", "Impact", "Summary", "Source"
    ]
    col_widths = [32, 15, 24, 12, 12, 55, 15]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # ── Data Rows ─────────────────────────────────────────────────
    if not announcements:
        ws.merge_cells("A3:G3")
        cell = ws["A3"]
        cell.value = "No other announcements found."
        cell.font = Font(name="Calibri", size=11, italic=True, color="888888")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A3"
        return

    for sr_no, ann in enumerate(announcements, 1):
        ai_data = ann.get("ai_data", {}) or {}
        row = sr_no + 2

        try:
            ann_date = datetime.fromisoformat(ann.get("announcement_date", "")).strftime("%d.%m.%Y")
        except:
            ann_date = ""

        sentiment = ai_data.get("sentiment", "Neutral")
        summary = ai_data.get("description") or ai_data.get("key_details") or ann.get("raw_subject", "")

        row_data = [
            ai_data.get("company_name") or ann.get("company_name", ""),
            ann_date,
            ai_data.get("announcement_type", "Other"),
            sentiment,
            ai_data.get("impact_level") or ai_data.get("impact", "Low"),
            summary,
            ann.get("source_url", ""),
        ]

        is_alt = (sr_no % 2 == 0)
        if sentiment == "Positive":
            row_fill = PatternFill("solid", fgColor=COLORS["positive_bg"])
        elif sentiment == "Negative":
            row_fill = PatternFill("solid", fgColor=COLORS["negative_bg"])
        elif is_alt:
            row_fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        else:
            row_fill = None

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            if col_idx in [1, 6]:
                align = "left"
            else:
                align = "center"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row].height = 35

    ws.freeze_panes = "A3"


def _save_workbook_bytes(wb: Workbook) -> bytes:
    """Save workbook to bytes buffer."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
